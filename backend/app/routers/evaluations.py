# backend/app/routers/evaluations.py
# Endpoints CRUD para evaluaciones físicas y exportación de reportes

from fastapi import APIRouter, Depends, HTTPException, status, Response
from sqlalchemy.orm import Session
from typing import List, Optional
import logging
from datetime import date

from app.database import get_db
from app.models.user import User
from app.models.patient import Patient
from app.models.evaluation import Evaluation
from app.schemas import EvaluationCreate, EvaluationResponse
from app.utils.security import get_current_active_user
from app.utils.calculations import (
    calcular_imc, calcular_indice_ruffier,
    generar_alertas, comparar_evaluaciones
)

# -----------------------------------------------
# Configuración del router de evaluaciones
# -----------------------------------------------
router = APIRouter(prefix="/evaluations", tags=["Evaluaciones"])
logger = logging.getLogger(__name__)


@router.post("/patients/{patient_id}", response_model=EvaluationResponse, status_code=status.HTTP_201_CREATED)
async def crear_evaluacion(
    patient_id: int,
    eval_data: EvaluationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Registra una nueva evaluación física para un paciente.
    Calcula automáticamente IMC, Índice Ruffier y genera alertas.
    """
    # Verificar que el paciente pertenece al entrenador
    paciente = _verificar_paciente(patient_id, current_user.id, db)

    # Calcular número de evaluación secuencial
    num_evaluacion = db.query(Evaluation).filter(
        Evaluation.patient_id == patient_id
    ).count() + 1

    # Preparar datos de la evaluación
    eval_dict = eval_data.model_dump(exclude_none=True)

    # Calcular IMC automáticamente si hay peso y talla
    peso = eval_dict.get("peso_kg") or paciente.peso_inicial_kg
    talla = eval_dict.get("talla_metros") or paciente.talla_metros
    imc_calculado = calcular_imc(peso, talla)

    # Calcular Índice de Ruffier si hay las tres frecuencias cardíacas
    indice_ruffier = None
    if all(k in eval_dict for k in ["fc_reposo", "fc_post_esfuerzo", "fc_minuto_recuperacion"]):
        indice_ruffier = calcular_indice_ruffier(
            eval_dict["fc_reposo"],
            eval_dict["fc_post_esfuerzo"],
            eval_dict["fc_minuto_recuperacion"]
        )

    # Generar alertas de salud automáticas
    eval_con_imc = {**eval_dict, "imc": imc_calculado}
    tiene_alerta, detalle_alerta = generar_alertas(eval_con_imc)

    # Crear el registro de evaluación
    nueva_evaluacion = Evaluation(
        patient_id=patient_id,
        trainer_id=current_user.id,
        numero_evaluacion=num_evaluacion,
        imc=imc_calculado,
        indice_ruffier=indice_ruffier,
        tiene_alerta=tiene_alerta,
        detalle_alerta=detalle_alerta,
        **eval_dict
    )

    db.add(nueva_evaluacion)
    db.commit()
    db.refresh(nueva_evaluacion)

    logger.info(f"Evaluación creada: {nueva_evaluacion.id} para paciente: {patient_id}")

    return nueva_evaluacion


@router.get("/patients/{patient_id}", response_model=List[EvaluationResponse])
async def listar_evaluaciones_paciente(
    patient_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Lista todas las evaluaciones históricas de un paciente.
    Ordenadas cronológicamente ascendente para análisis de progreso.
    """
    # Verificar acceso al paciente
    _verificar_paciente(patient_id, current_user.id, db)

    evaluaciones = db.query(Evaluation).filter(
        Evaluation.patient_id == patient_id
    ).order_by(Evaluation.fecha_evaluacion.asc()).all()

    return evaluaciones


@router.get("/{evaluation_id}", response_model=EvaluationResponse)
async def obtener_evaluacion(
    evaluation_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Obtiene los datos completos de una evaluación específica.
    Verifica que pertenezca al entrenador autenticado.
    """
    evaluacion = _get_evaluation_or_404(evaluation_id, current_user.id, db)
    return evaluacion


@router.put("/{evaluation_id}", response_model=EvaluationResponse)
async def actualizar_evaluacion(
    evaluation_id: int,
    eval_data: EvaluationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Actualiza una evaluación existente y recalcula indicadores automáticos.
    Recalcula IMC, Ruffier y alertas con los nuevos datos.
    """
    evaluacion = _get_evaluation_or_404(evaluation_id, current_user.id, db)

    # Actualizar campos con los nuevos datos
    update_data = eval_data.model_dump(exclude_unset=True)
    for campo, valor in update_data.items():
        setattr(evaluacion, campo, valor)

    # Recalcular IMC con nuevos valores
    if evaluacion.peso_kg and evaluacion.talla_metros:
        evaluacion.imc = calcular_imc(evaluacion.peso_kg, evaluacion.talla_metros)

    # Recalcular Ruffier si están los datos
    if all([evaluacion.fc_reposo, evaluacion.fc_post_esfuerzo, evaluacion.fc_minuto_recuperacion]):
        evaluacion.indice_ruffier = calcular_indice_ruffier(
            evaluacion.fc_reposo, evaluacion.fc_post_esfuerzo, evaluacion.fc_minuto_recuperacion
        )

    # Regenerar alertas con datos actualizados
    eval_dict = {c.name: getattr(evaluacion, c.name) for c in evaluacion.__table__.columns}
    evaluacion.tiene_alerta, evaluacion.detalle_alerta = generar_alertas(eval_dict)

    db.commit()
    db.refresh(evaluacion)

    return evaluacion


@router.delete("/{evaluation_id}", status_code=status.HTTP_204_NO_CONTENT)
async def eliminar_evaluacion(
    evaluation_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Elimina permanentemente una evaluación.
    Operación destructiva - requiere confirmación en el frontend.
    """
    evaluacion = _get_evaluation_or_404(evaluation_id, current_user.id, db)

    db.delete(evaluacion)
    db.commit()

    logger.info(f"Evaluación eliminada: {evaluation_id}")


@router.get("/patients/{patient_id}/comparar")
async def comparar_evaluaciones_paciente(
    patient_id: int,
    eval_id_1: int,
    eval_id_2: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Compara dos evaluaciones del mismo paciente y retorna los deltas.
    Indica si cada indicador mejoró, empeoró o se mantuvo.
    """
    _verificar_paciente(patient_id, current_user.id, db)

    # Obtener las dos evaluaciones a comparar
    eval1 = db.query(Evaluation).filter(
        Evaluation.id == eval_id_1,
        Evaluation.patient_id == patient_id
    ).first()

    eval2 = db.query(Evaluation).filter(
        Evaluation.id == eval_id_2,
        Evaluation.patient_id == patient_id
    ).first()

    if not eval1 or not eval2:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Una o ambas evaluaciones no encontradas"
        )

    # Convertir a diccionarios para comparación
    def eval_to_dict(e):
        return {c.name: getattr(e, c.name) for c in e.__table__.columns}

    comparacion = comparar_evaluaciones(eval_to_dict(eval1), eval_to_dict(eval2))

    return {
        "evaluacion_anterior": EvaluationResponse.model_validate(eval1),
        "evaluacion_actual": EvaluationResponse.model_validate(eval2),
        "comparacion": comparacion
    }


@router.get("/patients/{patient_id}/export/excel")
async def exportar_excel(
    patient_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Genera y descarga un archivo Excel con el historial completo del paciente.
    Incluye hoja de datos del paciente y hoja de evaluaciones.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    import openpyxl.utils

    paciente = _verificar_paciente(patient_id, current_user.id, db)
    evaluaciones = db.query(Evaluation).filter(
        Evaluation.patient_id == patient_id
    ).order_by(Evaluation.fecha_evaluacion.asc()).all()

    # Crear libro de trabajo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluaciones"

    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # Encabezados de la tabla de evaluaciones
    encabezados = [
        "Fecha", "Evaluación #", "Peso (kg)", "IMC", "% Grasa", "% Agua",
        "Músculo (kg)", "Oxigenación", "FC (lpm)", "TA", "Perímetro Abd.",
        "Índice Ruffier", "Fuerza Der. (kg)", "Test Wells (cm)",
        "Condición", "Alertas"
    ]

    for col, header in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Insertar datos de cada evaluación
    for row, eva in enumerate(evaluaciones, 2):
        ws.cell(row=row, column=1, value=str(eva.fecha_evaluacion))
        ws.cell(row=row, column=2, value=eva.numero_evaluacion)
        ws.cell(row=row, column=3, value=eva.peso_kg)
        ws.cell(row=row, column=4, value=eva.imc)
        ws.cell(row=row, column=5, value=eva.porcentaje_grasa)
        ws.cell(row=row, column=6, value=eva.porcentaje_agua)
        ws.cell(row=row, column=7, value=eva.musculo_kg)
        ws.cell(row=row, column=8, value=eva.oxigenacion_porcentaje)
        ws.cell(row=row, column=9, value=eva.frecuencia_cardiaca_rpm)
        ws.cell(row=row, column=10, value=f"{eva.tension_sistolica}/{eva.tension_diastolica}" if eva.tension_sistolica else "")
        ws.cell(row=row, column=11, value=eva.perimetro_abdominal_cm)
        ws.cell(row=row, column=12, value=eva.indice_ruffier)
        ws.cell(row=row, column=13, value=eva.fuerza_manual_der_kg)
        ws.cell(row=row, column=14, value=eva.test_wells_cm)
        ws.cell(row=row, column=15, value=eva.condicion_fisica.value if eva.condicion_fisica else "")
        ws.cell(row=row, column=16, value=eva.detalle_alerta or "Sin alertas")

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = min(max_length + 4, 40)

    # Guardar en buffer de memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"fitpro_{paciente.nombre_completo.replace(' ', '_')}_{date.today()}.xlsx"

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )


def _verificar_paciente(patient_id: int, trainer_id: int, db: Session) -> Patient:
    """Verifica que el paciente existe y pertenece al entrenador"""
    paciente = db.query(Patient).filter(
        Patient.id == patient_id,
        Patient.trainer_id == trainer_id,
        Patient.is_active == True
    ).first()

    if not paciente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Paciente no encontrado o sin acceso"
        )
    return paciente


def _get_evaluation_or_404(eval_id: int, trainer_id: int, db: Session) -> Evaluation:
    """Verifica que la evaluación existe y pertenece al entrenador"""
    evaluacion = db.query(Evaluation).filter(
        Evaluation.id == eval_id,
        Evaluation.trainer_id == trainer_id
    ).first()

    if not evaluacion:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evaluación no encontrada"
        )
    return evaluacion